import csv

from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side


class Vacancy:

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.key_skills = vacancy['key_skills'].split('\n')
        self.year = int(vacancy['published_at'][:4])


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    def csv_reader(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            header_length = len(header)
            for row in reader:
                if len(row) == header_length and not row[1].strip() == '' and not row[1] == ' ':
                    t = row[1].replace(', ', '\n').replace('; ', '\n')
                    row[1] = t
                    yield dict(zip(header, row))

    def get_statistic(self):
        result = {}
        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            for a in self.vacancy_name.split(', '):
                if vacancy.name.find(a) != -1:
                    for skill in vacancy.key_skills:
                        if vacancy.year in result:
                            if skill in result[vacancy.year]:
                                result[vacancy.year][skill] += 1
                            else:
                                result[vacancy.year][skill] = 1
                        else:
                            result[vacancy.year] = {skill: 1}
                    break
        print(result)
        answer = {}
        for time in result:
            t = result[time]
            sorted_dic = dict(sorted(t.items(), key=lambda item: item[1], reverse=True))
            t = {}
            f = 0
            for skill in sorted_dic:
                t[skill] = sorted_dic[skill]
                f += 1
                if f == 11:
                    break
            answer[time] = t
        print(answer)
        return answer


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        result = dataset.get_statistic()
        for i in result:
            t = result[i]
            report = Report(result[i], i)
            report.generate_excel()
            t.pop(1)
            report.generate_image(result[i])


class Report:
    def __init__(self, result, n):
        self.wb = Workbook()
        self.result = result
        self.n = n

    def generate_excel(self):
        ws1 = self.wb.active
        ws1.title = 'Топ-10 навыков {} года'.format(str(self.n))
        ws1.append(['skill', 'count'])
        for year in self.result:
            ws1.append([year, self.result[year]])

        data = [['skill', 'count']]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'AB':
            ws1[col + '1'].font = font_bold

        thin = Side(border_style='thin', color='00000000')

        self.result[1] = 1
        for row, _ in enumerate(self.result):
            for col in 'AB':
                ws1[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.wb.save('Навыки по годам\\report_backend{}.xlsx'.format(str(self.n)))

    def generate_image(self, result):
        fig, ax = plt.subplots()
        my_labels = []
        values = []
        sum = 0
        for i in result:
            my_labels.append(i)
            sum += result[i]
            values.append(result[i])
        for i in range(0, len(values)):
            values[i] = float(values[i] / sum)
        ax.set_title('Доля навыков по годам', fontdict={'fontsize': 15})
        ax.pie(values, labels=my_labels)
        plt.legend(title="Топ-10 навыков:")
        plt.tight_layout()
        plt.savefig('Навыки по годам\\graph{}.png'.format(str(self.n)))


if __name__ == '__main__':
    InputConnect()
